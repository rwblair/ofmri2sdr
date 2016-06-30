from openpyxl import Workbook, load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

wb = load_workbook('ofmri.citation.xlsx')
ws  = wb.active
out_wb = Workbook(write_only=True)
out_ws = out_wb.create_sheet()

for row in range(2,51):
    names = ""
    out = []
    for col in range(2, 33):
        value = ws.cell(column=col, row=row).value
        if value and value != "Principal investigator":
            names += value
            names += '; '
        if not value:
            # remove last comma
            index = names.rfind(';')
            names = list(names)
            names[index] = '.'
            names = ''.join(names)
            break
    date = ws.cell(column=34, row=row).value
    try:
        date = str(date.year)
    except AttributeError:
        date = str(date)[:4]
    title = ws.cell(column=1, row=row).value
    citation = names + "(" + date + "). " + title + ". Stanford Digital Repository. Available at: http://purl.stanford.edu/[DRUID] and https://openfmri.org/dataset/"
    out.append(citation)
    out_ws.append(out)

out_wb.save(filename='citation.xlsx')
